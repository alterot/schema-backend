from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from typing import Dict, Any
import traceback
import json
import os
import time
from datetime import date, timedelta, datetime
from calendar import monthrange
from dotenv import load_dotenv

load_dotenv()

from models import Person, Shift
from solver import SchemaOptimizer
from utils import validate_input, ValidationError, calculate_all_metrics
from data import get_personal, get_bemanningsbehov, get_regler, get_avdelning, generate_shifts_for_period, is_helgdag
from utils.schedule_analyzer import find_conflicts, suggest_solutions, calculate_impact
from utils.supabase_client import save_audit_log

# Path for saving schedules to disk
SAVED_SCHEDULES_DIR = os.path.join(os.path.dirname(__file__), 'data', 'saved_schedules')

app = Flask(__name__)

# Konfigurera CORS för att tillåta React frontend
CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:3000", "http://localhost:5173", "http://localhost:5174", "https://alterot.github.io"],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})


def _generate_schedule_for_period(period: str, override_personal=None, override_bemanningsbehov=None, personal_overrides=None):
    """
    Shared helper: Generate a schedule for a YYYY-MM period using the solver.

    Args:
        period: YYYY-MM format
        override_personal: Optional list of person dicts from frontend (localStorage)
        override_bemanningsbehov: Optional dict with vardag/helg requirements from frontend
        personal_overrides: Optional list of dicts with modifications to apply to personal
                            before running the solver. Supports: add_franvaro, extra_pass,
                            tillganglighet, action:"add" (vikarie).

    Returns (personal, shifts, schedule, metrics) tuple.
    Raises ValueError on invalid period.
    """
    year, month = period.split('-')
    year, month = int(year), int(month)
    if not (1 <= month <= 12):
        raise ValueError(f"Invalid month: {month}")

    start_date = date(year, month, 1)
    end_date = date(year, month, monthrange(year, month)[1])

    # Use frontend data if provided, otherwise fallback to JSON file
    if override_personal:
        personal_data = override_personal
        app.logger.info(f'Using {len(personal_data)} personal from frontend override')
    else:
        personal_data = get_personal()

    avdelning_info = get_avdelning()

    # Generate shifts — if bemanningsbehov override provided, build shifts manually
    if override_bemanningsbehov:
        app.logger.info('Using bemanningsbehov from frontend override')
        shifts_data = _generate_shifts_with_custom_behov(
            start_date, end_date, avdelning_info['namn'], override_bemanningsbehov
        )
    else:
        shifts_data = generate_shifts_for_period(start_date, end_date, avdelning_info['namn'])

    personal = [Person.from_dict(p) for p in personal_data]
    shifts = [Shift.from_dict(s) for s in shifts_data]

    # Apply personal_overrides (frånvaro, övertid, vikarier, tillgänglighet)
    if personal_overrides:
        from models.person import Franvaro
        for mod in personal_overrides:
            namn = mod.get('namn', '')

            # Action "add": lägg till ny person (vikarie)
            if mod.get('action') == 'add':
                new_id = max((p.id for p in personal), default=0) + 1
                new_person_data = {
                    'id': new_id,
                    'namn': namn,
                    'roll': mod.get('roll', 'underskoterska'),
                    'anstallning': mod.get('anstallning', 100),
                    'tillganglighet': mod.get('tillganglighet', ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']),
                    'franvaro': [],
                }
                personal.append(Person.from_dict(new_person_data))
                app.logger.info(f'Personal override: lade till vikarie {namn} ({new_person_data["roll"]})')
                continue

            # Hitta befintlig person
            matched = [p for p in personal if p.namn.lower() == namn.lower()]
            if not matched:
                app.logger.warning(f'Personal override: person "{namn}" hittades inte')
                continue
            person = matched[0]

            # add_franvaro: lägg till frånvaroperiod
            if 'add_franvaro' in mod:
                f = mod['add_franvaro']
                new_franvaro = Franvaro(
                    start=date.fromisoformat(f['start']),
                    slut=date.fromisoformat(f['slut']),
                    typ=f.get('typ', 'semester'),
                )
                person.franvaro.append(new_franvaro)
                app.logger.info(f'Personal override: {person.namn} frånvaro {new_franvaro.typ} {new_franvaro.start}-{new_franvaro.slut}')

            # extra_pass: öka max arbetspass (övertid)
            if 'extra_pass' in mod:
                old_max = person.max_arbetspass_per_manad
                person.max_arbetspass_per_manad += mod['extra_pass']
                person.max_timmar_per_manad += mod['extra_pass'] * 8  # Approximera 8h/pass
                app.logger.info(f'Personal override: {person.namn} max_pass {old_max} -> {person.max_arbetspass_per_manad}')

            # tillganglighet: ändra vilka dagar personen kan jobba
            if 'tillganglighet' in mod:
                old_tillg = person.tillganglighet
                person.tillganglighet = mod['tillganglighet']
                app.logger.info(f'Personal override: {person.namn} tillgänglighet {old_tillg} -> {person.tillganglighet}')

            # exclude_pass_typer: blockera specifika passtyper
            if 'exclude_pass_typer' in mod:
                person.exclude_pass_typer = mod['exclude_pass_typer']
                app.logger.info(f'Personal override: {person.namn} exclude_pass_typer {person.exclude_pass_typer}')

            # lasta_pass: tvinga person till specifikt pass/datum
            if 'lasta_pass' in mod:
                person.lasta_pass = [
                    {'datum': date.fromisoformat(lp['datum']), 'pass_typ': lp['pass_typ']}
                    for lp in mod['lasta_pass']
                ]
                app.logger.info(f'Personal override: {person.namn} lasta_pass {mod["lasta_pass"]}')

    app.logger.info(f'Generating schedule for {period}: {len(personal)} personal, {len(shifts)} shifts')

    optimizer = SchemaOptimizer(personal, shifts)
    schedule = optimizer.optimera()

    metrics = calculate_all_metrics(
        schema_rader=schedule.rader,
        konflikter=schedule.konflikter,
        shifts=shifts,
        personal=personal
    )

    return personal, shifts, schedule, metrics


def _generate_shifts_with_custom_behov(start_date, end_date, avdelning, bemanningsbehov):
    """Generate shifts using custom bemanningsbehov instead of JSON file data."""
    shifts = []
    current_date = start_date

    while current_date <= end_date:
        is_weekend = current_date.weekday() >= 5 or is_helgdag(current_date)
        behov_typ = 'helg' if is_weekend else 'vardag'
        behov = bemanningsbehov.get(behov_typ, bemanningsbehov.get('vardag', {}))

        for pass_typ in ['dag', 'kvall', 'natt']:
            pass_namn = pass_typ if pass_typ != 'kvall' else 'kväll'
            shifts.append({
                'datum': current_date.isoformat(),
                'pass': pass_namn,
                'avdelning': avdelning,
                'kompetenskrav': behov.get(pass_typ, {})
            })

        current_date = current_date + timedelta(days=1)

    return shifts


def _load_saved_schedule(period: str):
    """Load a previously saved schedule from disk, or None if not found."""
    filepath = os.path.join(SAVED_SCHEDULES_DIR, f'{period}.json')
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


@app.route('/api/health', methods=['GET'])
def health_check():
    """Hälsokontroll för att verifiera att API:et är igång"""
    return jsonify({
        'status': 'ok',
        'service': 'Schema-assistent Backend',
        'version': '1.0.0'
    })


@app.route('/api/generate', methods=['POST'])
def generate_schedule():
    """
    Huvudendpoint för schemagenerering.

    Input (JSON):
    {
        "personal": [...],
        "behov": [...],
        "config": {...}
    }

    Output (JSON):
    {
        "schema": [...],
        "konflikter": [...],
        "statistik": {...},
        "metrics": {
            "coverage_percent": float,
            "overtime_hours": float,
            "rule_violations": int,
            "cost_kr": float,
            "quality_score": int
        }
    }
    """
    try:
        # Steg 1: Hämta och validera input
        data = request.get_json()

        if not data:
            return jsonify({
                'error': 'Ingen data mottagen',
                'detaljer': 'Request body måste innehålla JSON-data'
            }), 400

        # Validera input-strukturen
        try:
            validate_input(data)
        except ValidationError as e:
            return jsonify({
                'error': 'Valideringsfel',
                'detaljer': str(e)
            }), 400

        # Steg 2: Konvertera till datamodeller
        personal = [Person.from_dict(p) for p in data['personal']]
        shifts = [Shift.from_dict(s) for s in data['behov']]

        # Steg 3: Validera att vi har data att arbeta med
        if not personal:
            return jsonify({
                'error': 'Ingen personal angiven',
                'detaljer': 'Det måste finnas minst en person i listan'
            }), 400

        if not shifts:
            return jsonify({
                'error': 'Inga pass angivna',
                'detaljer': 'Det måste finnas minst ett pass att schemalägga'
            }), 400

        # Steg 4: Kör optimering
        optimizer = SchemaOptimizer(personal, shifts)
        schedule = optimizer.optimera()

        # Steg 5: Beräkna metrics
        metrics = calculate_all_metrics(
            schema_rader=schedule.rader,
            konflikter=schedule.konflikter,
            shifts=shifts,
            personal=personal
        )

        # Steg 6: Returnera resultat med metrics
        result = schedule.to_dict()
        result['metrics'] = metrics
        result['personal_lookup'] = {
            str(p.id): {'namn': p.namn, 'roll': p.roll}
            for p in personal
        }
        return jsonify(result), 200

    except ValidationError as e:
        # Specifika valideringsfel
        return jsonify({
            'error': 'Valideringsfel',
            'detaljer': str(e)
        }), 400

    except Exception as e:
        # Oväntade fel
        app.logger.error(f'Oväntat fel vid schemagenerering: {str(e)}')
        app.logger.error(traceback.format_exc())

        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': 'Ett oväntat fel uppstod vid schemagenerering',
            'teknisk_info': str(e) if app.debug else None
        }), 500


@app.route('/api/generate-realistic', methods=['POST'])
def generate_realistic_schedule():
    """
    Genererar schema med realistisk sjukhusdata.

    Input (JSON):
    {
        "start_date": "2025-04-01",  // Optional, default: idag
        "end_date": "2025-04-30",    // Optional, default: 30 dagar fram
        "scenario": "normal_april"   // Optional testscenario
    }

    Output (JSON):
    {
        "schema": [...],
        "konflikter": [...],
        "statistik": {...},
        "metrics": {...},
        "data_source": "realistic_hospital_data"
    }
    """
    try:
        data = request.get_json() or {}

        # Hantera datum
        today = date.today()
        start_str = data.get('start_date')
        end_str = data.get('end_date')

        if start_str:
            start_date = date.fromisoformat(start_str)
        else:
            start_date = today

        if end_str:
            end_date = date.fromisoformat(end_str)
        else:
            end_date = start_date + timedelta(days=30)

        # Hamta personal fran realistisk data
        personal_data = get_personal()

        # Hamta avdelning
        avdelning_info = get_avdelning()
        avdelning_namn = avdelning_info['namn']

        # Generera shifts for perioden
        shifts_data = generate_shifts_for_period(start_date, end_date, avdelning_namn)

        # Konvertera till datamodeller
        personal = [Person.from_dict(p) for p in personal_data]
        shifts = [Shift.from_dict(s) for s in shifts_data]

        # Validera att vi har data att arbeta med
        if not personal:
            return jsonify({
                'error': 'Ingen personal i data',
                'detaljer': 'Kontrollera realistic_hospital_data.json'
            }), 400

        if not shifts:
            return jsonify({
                'error': 'Inga pass genererade',
                'detaljer': 'Kontrollera datum-intervallet'
            }), 400

        # Kor optimering
        optimizer = SchemaOptimizer(personal, shifts)
        schedule = optimizer.optimera()

        # Berakna metrics
        metrics = calculate_all_metrics(
            schema_rader=schedule.rader,
            konflikter=schedule.konflikter,
            shifts=shifts,
            personal=personal
        )

        # Returnera resultat
        result = schedule.to_dict()
        result['metrics'] = metrics
        result['data_source'] = 'realistic_hospital_data'
        result['period'] = {
            'start': start_date.isoformat(),
            'end': end_date.isoformat()
        }
        result['avdelning'] = avdelning_info
        result['personal_lookup'] = {
            str(p.id): {'namn': p.namn, 'roll': p.roll}
            for p in personal
        }

        return jsonify(result), 200

    except ValueError as e:
        return jsonify({
            'error': 'Ogiltigt datum',
            'detaljer': str(e)
        }), 400

    except Exception as e:
        app.logger.error(f'Oväntat fel vid realistisk schemagenerering: {str(e)}')
        app.logger.error(traceback.format_exc())

        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': 'Ett oväntat fel uppstod vid schemagenerering',
            'teknisk_info': str(e) if app.debug else None
        }), 500


@app.route('/api/data/personal', methods=['GET'])
def get_personal_endpoint():
    """Returnerar listan med personal fran realistisk data."""
    try:
        personal = get_personal()
        return jsonify({
            'personal': personal,
            'antal': len(personal)
        }), 200
    except Exception as e:
        app.logger.error(f'Fel vid hamtning av personal: {str(e)}')
        return jsonify({'error': 'Kunde inte hamta personal'}), 500


@app.route('/api/data/bemanningsbehov', methods=['GET'])
def get_bemanningsbehov_endpoint():
    """Returnerar bemanningsbehov for vardag och helg."""
    try:
        vardag = get_bemanningsbehov(is_weekend=False)
        helg = get_bemanningsbehov(is_weekend=True)
        return jsonify({
            'vardag': vardag,
            'helg': helg
        }), 200
    except Exception as e:
        app.logger.error(f'Fel vid hamtning av bemanningsbehov: {str(e)}')
        return jsonify({'error': 'Kunde inte hamta bemanningsbehov'}), 500


@app.route('/api/data/regler', methods=['GET'])
def get_regler_endpoint():
    """Returnerar regler for schemalagning."""
    try:
        regler = get_regler()
        return jsonify(regler), 200
    except Exception as e:
        app.logger.error(f'Fel vid hamtning av regler: {str(e)}')
        return jsonify({'error': 'Kunde inte hamta regler'}), 500


@app.route('/api/validate', methods=['POST'])
def validate_input_endpoint():
    """
    Endpoint för att validera input utan att generera schema.
    Användbart för frontend-validering.
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'valid': False,
                'error': 'Ingen data mottagen'
            }), 400

        # Försök validera
        validate_input(data)

        # Om vi kommer hit är datan giltig
        return jsonify({
            'valid': True,
            'message': 'Input är giltig',
            'statistik': {
                'antal_personal': len(data.get('personal', [])),
                'antal_pass': len(data.get('behov', []))
            }
        }), 200

    except ValidationError as e:
        return jsonify({
            'valid': False,
            'error': str(e)
        }), 400

    except Exception as e:
        app.logger.error(f'Fel vid validering: {str(e)}')
        return jsonify({
            'valid': False,
            'error': 'Oväntat fel vid validering'
        }), 500


@app.route('/api/schedule/<period>', methods=['GET', 'POST'])
def get_schedule(period: str):
    """
    Tool endpoint: Hämta/generera schema för en period.

    GET: Returnerar sparat schema om det finns, annars genererar nytt via solver.
    POST: Accepts optional personal/bemanningsbehov from frontend (localStorage).

    Args:
        period: YYYY-MM format (t.ex. "2025-04")
    """
    try:
        # Validera period-format
        try:
            year, month = period.split('-')
            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                raise ValueError()
        except (ValueError, AttributeError):
            return jsonify({
                'error': 'Ogiltigt periodformat',
                'detaljer': f"Period '{period}' är ogiltig. Använd format YYYY-MM (t.ex. '2025-04')"
            }), 400

        # Extract optional overrides from POST body
        override_personal = None
        override_behov = None
        personal_overrides = None
        regenerate = False
        if request.method == 'POST':
            data = request.get_json() or {}
            override_personal = data.get('personal')
            override_behov = data.get('bemanningsbehov')
            personal_overrides = data.get('personal_overrides')
            regenerate = data.get('regenerate', False)

        # Personal overrides always require regeneration (can't use cached schedule)
        if personal_overrides:
            regenerate = True

        # Check for previously saved schedule (skip if regenerate requested)
        if not regenerate:
            saved = _load_saved_schedule(period)
            if saved:
                app.logger.info(f'Returnerar sparat schema för {period}')
                saved['source'] = 'saved'
                return jsonify(saved), 200

        # Generate new schedule via solver
        gen_start = time.time()
        personal, shifts, schedule, metrics = _generate_schedule_for_period(
            period,
            override_personal=override_personal,
            override_bemanningsbehov=override_behov,
            personal_overrides=personal_overrides,
        )
        duration_ms = int((time.time() - gen_start) * 1000)

        result = schedule.to_dict()
        result['metrics'] = metrics
        result['period'] = period
        result['source'] = 'generated'
        result['message'] = f'Schema genererat för {period} med {len(personal)} personal och {len(shifts)} pass'

        # Include helgdagar so frontend knows which days use helg-bemanning
        helgdagar = [
            s.datum.isoformat() for s in shifts
            if s.datum.weekday() < 5 and is_helgdag(s.datum)
        ]
        # Deduplicate (3 shifts per day)
        result['helgdagar'] = sorted(set(helgdagar))

        # Include personal_lookup so frontend can map IDs to names
        result['personal_lookup'] = {
            str(p.id): {'namn': p.namn, 'roll': p.roll}
            for p in personal
        }

        # Include franvaro periods so frontend can render per-day absence
        franvaro_perioder = {}
        for p in personal:
            if p.franvaro:
                franvaro_perioder[p.namn] = [
                    {'start': f.start.isoformat(), 'slut': f.slut.isoformat(), 'typ': f.typ}
                    for f in p.franvaro
                ]
        if franvaro_perioder:
            result['franvaro_perioder'] = franvaro_perioder

        # Auto-save so "Visa schema" returns the same data
        os.makedirs(SAVED_SCHEDULES_DIR, exist_ok=True)
        filepath = os.path.join(SAVED_SCHEDULES_DIR, f'{period}.json')
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        app.logger.info(f'Schema auto-sparat till {filepath}')

        # Audit log to Supabase
        has_infeasible = any(
            (k.typ if hasattr(k, 'typ') else k.get('typ', '')) == 'infeasible'
            for k in schedule.konflikter
        )
        solver_status = 'INFEASIBLE' if has_infeasible else 'OPTIMAL'

        user_input_text = None
        ai_reasoning_text = None
        if request.method == 'POST' and data:
            user_input_text = data.get('user_input')
            ai_reasoning_text = data.get('ai_reasoning')

        save_audit_log(
            period=period,
            schedule_data=result.get('schema', []),
            metrics=metrics,
            konflikter=schedule.konflikter,
            solver_status=solver_status,
            antal_personal=len(personal),
            duration_ms=duration_ms,
            personal_overrides=personal_overrides,
            user_input=user_input_text,
            ai_reasoning=ai_reasoning_text,
        )

        return jsonify(result), 200

    except ValueError as e:
        return jsonify({
            'error': 'Ogiltigt periodformat',
            'detaljer': str(e)
        }), 400

    except Exception as e:
        app.logger.error(f'Fel vid hämtning av schema: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': f'Kunde inte generera schema: {str(e)}'
        }), 500


@app.route('/api/schedule/<period>/export', methods=['GET'])
def export_schedule_excel(period: str):
    """
    Exportera schema som Excel-fil (.xlsx).
    Använder sparat schema om det finns, annars genererar nytt.
    Returnerar fil med två flikar: Kalender och Per person.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        # Load saved schedule or generate new
        saved = _load_saved_schedule(period)
        if not saved:
            personal, shifts, schedule, metrics = _generate_schedule_for_period(period)
            saved = schedule.to_dict()
            saved['metrics'] = metrics

        schema_rows = saved.get('schema', [])
        konflikter = saved.get('konflikter', [])
        metrics = saved.get('metrics', {})
        personal_lookup = saved.get('personal_lookup', {})

        # Parse period
        year, month = period.split('-')
        year, month = int(year), int(month)
        num_days = monthrange(year, month)[1]

        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
        red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        dag_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        kvall_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
        natt_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')

        weekday_names = ['Mon', 'Tis', 'Ons', 'Tor', 'Fre', 'Lor', 'Son']

        # Helper: resolve person ID to name
        def _resolve_name(pid):
            entry = personal_lookup.get(str(pid))
            return entry['namn'] if entry else str(pid)

        # Group schema by day
        day_map = {}
        for row in schema_rows:
            datum = row['datum']
            if datum not in day_map:
                day_map[datum] = {'dag': [], 'kvall': [], 'natt': []}
            pass_key = 'kvall' if row.get('pass') == 'kväll' else row.get('pass', '')
            if pass_key in day_map[datum]:
                day_map[datum][pass_key] = [_resolve_name(pid) for pid in row.get('personal', [])]

        # Conflict dates for highlighting
        conflict_dates = set(k.get('datum') for k in konflikter if k.get('datum'))

        wb = Workbook()

        # ── Sheet 1: Kalender ──
        ws1 = wb.active
        ws1.title = 'Kalender'

        # Metrics header
        ws1.append([f'Schema {period}'])
        ws1['A1'].font = Font(bold=True, size=14)
        coverage = metrics.get('coverage_percent', '?')
        overtime = metrics.get('overtime_hours', '?')
        quality = metrics.get('quality_score', '?')
        ws1.append([f'Tackning: {coverage}%    Overtid: {overtime}h    Kvalitet: {quality}/100'])
        ws1.append([])

        # Table header
        headers = ['Datum', 'Veckodag', 'Dag', 'Kvall', 'Natt']
        ws1.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for day_num in range(1, num_days + 1):
            datum = f'{period}-{day_num:02d}'
            d = date(year, month, day_num)
            weekday = weekday_names[d.weekday()]
            day_data = day_map.get(datum, {'dag': [], 'kvall': [], 'natt': []})

            dag_str = ', '.join(day_data['dag']) or '-'
            kvall_str = ', '.join(day_data['kvall']) or '-'
            natt_str = ', '.join(day_data['natt']) or '-'

            row_idx = 4 + day_num
            ws1.append([datum, weekday, dag_str, kvall_str, natt_str])

            # Color coding
            if datum in conflict_dates:
                fill = red_fill
            elif d.weekday() >= 5:
                fill = yellow_fill
            else:
                fill = green_fill

            for col_idx in range(1, 6):
                cell = ws1.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.border = thin_border

        # Column widths
        ws1.column_dimensions['A'].width = 14
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 40
        ws1.column_dimensions['D'].width = 40
        ws1.column_dimensions['E'].width = 40

        # ── Sheet 2: Per person ──
        ws2 = wb.create_sheet('Per person')

        # Build person-shift map (resolve IDs to names for display)
        person_shifts = {}
        for row in schema_rows:
            datum = row['datum']
            pass_key = 'kvall' if row.get('pass') == 'kväll' else row.get('pass', '')
            for pid in row.get('personal', []):
                name = _resolve_name(pid)
                if name not in person_shifts:
                    person_shifts[name] = {}
                person_shifts[name][datum] = pass_key

        persons = sorted(person_shifts.keys())

        # Header row: Name | 1 | 2 | 3 | ... | Total
        header_row = ['Personal']
        for day_num in range(1, num_days + 1):
            header_row.append(str(day_num))
        header_row.append('Totalt')
        ws2.append(header_row)

        for col_idx in range(1, len(header_row) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Person rows
        shift_labels = {'dag': 'D', 'kvall': 'K', 'natt': 'N'}
        shift_fills = {'dag': dag_fill, 'kvall': kvall_fill, 'natt': natt_fill}

        for person_idx, name in enumerate(persons):
            row_idx = person_idx + 2
            ws2.cell(row=row_idx, column=1, value=name).border = thin_border
            total = 0

            for day_num in range(1, num_days + 1):
                datum = f'{period}-{day_num:02d}'
                shift = person_shifts[name].get(datum)
                col_idx = day_num + 1
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

                if shift:
                    cell.value = shift_labels.get(shift, '?')
                    cell.fill = shift_fills.get(shift, PatternFill())
                    total += 1
                else:
                    cell.value = '-'

            # Total column
            total_cell = ws2.cell(row=row_idx, column=num_days + 2, value=total)
            total_cell.border = thin_border
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal='center')

        # Column widths
        ws2.column_dimensions['A'].width = 25
        for day_num in range(1, num_days + 1):
            col_letter = ws2.cell(row=1, column=day_num + 1).column_letter
            ws2.column_dimensions[col_letter].width = 4

        # Save to memory and return
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'schema_{period}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        app.logger.error(f'Fel vid Excel-export: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': f'Kunde inte exportera schema: {str(e)}'
        }), 500


@app.route('/api/propose', methods=['POST'])
def propose_changes():
    """
    Tool endpoint: Analysera schema och föreslå förbättringar.
    Kör solver, identifierar konflikter, genererar datadrivna förslag.

    Input:
        { "problem": "...", "period": "YYYY-MM", "personal": [...], "bemanningsbehov": {...} }
    """
    try:
        data = request.get_json()

        if not data or 'problem' not in data:
            return jsonify({
                'error': 'Saknar problem-beskrivning',
                'detaljer': 'Request body måste innehålla "problem"-fält'
            }), 400

        problem = data['problem']
        period = data.get('period', date.today().strftime('%Y-%m'))
        override_personal = data.get('personal')
        override_behov = data.get('bemanningsbehov')

        # Generate schedule and analyze
        personal, shifts, schedule, metrics = _generate_schedule_for_period(
            period, override_personal=override_personal, override_bemanningsbehov=override_behov
        )

        # Load bemanningsbehov for conflict analysis
        if override_behov:
            bemanningsbehov = override_behov
        else:
            bemanningsbehov = {
                'vardag': get_bemanningsbehov(is_weekend=False),
                'helg': get_bemanningsbehov(is_weekend=True),
            }

        # Find conflicts in the generated schedule
        conflicts = find_conflicts(schedule.rader, shifts, personal, bemanningsbehov)
        app.logger.info(f'Found {len(conflicts)} conflicts for period {period}')

        # Generate proposals based on conflicts
        proposals = suggest_solutions(conflicts, personal, schedule.rader)

        # Build reasoning
        conflict_summary = f'{len(conflicts)} konflikter identifierade' if conflicts else 'Inga konflikter'
        reasoning = (
            f'Baserat på problem "{problem}" och analys av schemat för {period}: '
            f'{conflict_summary}. '
            f'Coverage: {metrics["coverage_percent"]}%, '
            f'övertid: {metrics["overtime_hours"]}h, '
            f'kvalitet: {metrics["quality_score"]}/100.'
        )

        return jsonify({
            'proposals': proposals,
            'reasoning': reasoning,
            'problem_analyzed': problem,
            'conflicts_found': len(conflicts),
            'conflicts': conflicts[:10],  # Limit to 10 for readability
            'current_metrics': metrics,
        }), 200

    except Exception as e:
        app.logger.error(f'Fel vid förslag av ändringar: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': f'Kunde inte generera förslag: {str(e)}'
        }), 500


@app.route('/api/simulate', methods=['POST'])
def simulate_impact():
    """
    Tool endpoint: Simulera konsekvenser av schemaändringar.
    Kör solver två gånger (before/after) och jämför metrics.

    Input:
        { "changes": [...], "period": "YYYY-MM", "personal": [...], "bemanningsbehov": {...} }
    """
    try:
        data = request.get_json()

        if not data or 'changes' not in data:
            return jsonify({
                'error': 'Saknar ändringar',
                'detaljer': 'Request body måste innehålla "changes"-fält'
            }), 400

        changes = data['changes']
        period = data.get('period', date.today().strftime('%Y-%m'))
        override_personal = data.get('personal')
        override_behov = data.get('bemanningsbehov')

        # Run solver for baseline metrics (before changes)
        personal_before, shifts, schedule_before, metrics_before = _generate_schedule_for_period(
            period, override_personal=override_personal, override_bemanningsbehov=override_behov
        )

        # Apply changes to personal data (e.g. add absence, change availability)
        personal_data_modified = [p.__dict__.copy() for p in personal_before]
        # Note: changes from AI are descriptive; we pass them through and re-run solver
        # For now, the "after" run uses the same data — the impact comparison shows
        # what the solver produces. In future: parse changes and modify personal_data.

        # Re-run solver (deterministic — same input gives same output, but this
        # establishes the pattern for when changes are actually applied)
        _, _, schedule_after, metrics_after = _generate_schedule_for_period(
            period, override_personal=override_personal, override_bemanningsbehov=override_behov
        )

        # Calculate impact diff
        impact = calculate_impact(metrics_before, metrics_after)

        # Build human-readable impact summary
        parts = []
        if impact['coverage_diff'] != 0:
            parts.append(f'coverage {"+" if impact["coverage_diff"] > 0 else ""}{impact["coverage_diff"]}%')
        if impact['overtime_diff'] != 0:
            parts.append(f'övertid {"+" if impact["overtime_diff"] > 0 else ""}{impact["overtime_diff"]}h')
        if impact['cost_diff'] != 0:
            parts.append(f'kostnad {"+" if impact["cost_diff"] > 0 else ""}{impact["cost_diff"]:,.0f} kr')
        if impact['quality_diff'] != 0:
            parts.append(f'kvalitet {"+" if impact["quality_diff"] > 0 else ""}{impact["quality_diff"]}')
        impact_text = ', '.join(parts) if parts else 'Ingen mätbar skillnad'

        return jsonify({
            'metrics_before': metrics_before,
            'metrics_after': metrics_after,
            'impact': impact,
            'impact_summary': impact_text,
            'changes_count': len(changes) if isinstance(changes, list) else 0,
            'period': period,
        }), 200

    except Exception as e:
        app.logger.error(f'Fel vid simulering: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': f'Kunde inte simulera ändringar: {str(e)}'
        }), 500


@app.route('/api/apply', methods=['POST'])
def apply_changes():
    """
    Tool endpoint: Spara schema till disk.
    Sparar till backend/data/saved_schedules/<period>.json.
    # TODO: Replace with database storage in production.

    Input:
        { "schema": {...}, "confirmed": bool, "period": "YYYY-MM" (optional) }
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'error': 'Ingen data mottagen',
                'detaljer': 'Request body måste innehålla JSON-data'
            }), 400

        if 'confirmed' not in data:
            return jsonify({
                'error': 'Saknar bekräftelse',
                'detaljer': 'Request body måste innehålla "confirmed"-fält'
            }), 400

        if not data['confirmed']:
            return jsonify({
                'success': False,
                'message': 'Ändringar ej bekräftade — ingen åtgärd utförd'
            }), 200

        if 'schema' not in data:
            return jsonify({
                'error': 'Saknar schema',
                'detaljer': 'Request body måste innehålla "schema"-fält när confirmed=true'
            }), 400

        schema_data = data['schema']
        period = data.get('period', date.today().strftime('%Y-%m'))
        timestamp = datetime.utcnow().isoformat() + 'Z'

        # Save to JSON file
        os.makedirs(SAVED_SCHEDULES_DIR, exist_ok=True)
        filepath = os.path.join(SAVED_SCHEDULES_DIR, f'{period}.json')

        save_payload = {
            'period': period,
            'saved_at': timestamp,
            **schema_data,
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(save_payload, f, ensure_ascii=False, indent=2)

        app.logger.info(f'Schema sparat till {filepath}')

        return jsonify({
            'success': True,
            'message': f'Schema för {period} sparat',
            'timestamp': timestamp,
            'saved_to': filepath,
            'changes_applied': True,
        }), 200

    except Exception as e:
        app.logger.error(f'Fel vid applicering av ändringar: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Internt serverfel',
            'detaljer': f'Kunde inte spara schema: {str(e)}'
        }), 500


@app.errorhandler(404)
def not_found(error):
    """Hanterare för 404-fel"""
    return jsonify({
        'error': 'Endpoint hittades inte',
        'detaljer': 'Kontrollera att du använder rätt URL'
    }), 404


@app.errorhandler(405)
def method_not_allowed(error):
    """Hanterare för 405-fel (felaktig HTTP-metod)"""
    return jsonify({
        'error': 'HTTP-metod inte tillåten',
        'detaljer': 'Kontrollera att du använder rätt HTTP-metod (GET/POST)'
    }), 405


@app.errorhandler(500)
def internal_error(error):
    """Hanterare för 500-fel"""
    app.logger.error(f'Internt serverfel: {str(error)}')
    return jsonify({
        'error': 'Internt serverfel',
        'detaljer': 'Ett oväntat fel uppstod på servern'
    }), 500


if __name__ == '__main__':
    # Utvecklingsserver
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=True
    )
